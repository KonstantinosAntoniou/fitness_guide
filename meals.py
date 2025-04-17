import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from foods import Food


class Meal(Food):
    def __init__(self, meal_name):
        super().__init__()  # No need to pass food-specific attributes
        self.meal_name = meal_name
        self.meal_excel_file = 'meals_log.xlsx'

        # Ensure the meals Excel file exists
        if not os.path.exists(self.meal_excel_file):
            self.create_meal_excel_file()

    def create_meal_excel_file(self):
        columns = ['Meal_Name', 'Food_Name', 'Label', 'Measurement', 'Calories', 'Protein', 'Carbs', 'Fat_Saturated', 'Fat_Regular', 'Sodium']
        df = pd.DataFrame(columns=columns)
        df.to_excel(self.meal_excel_file, index=False)

    def add_food_to_meal(self, food_name, food_label, multiplier=1):
        # Load the food Excel file
        df_food = pd.read_excel(self.excel_file)

        # Check if the food is logged
        food_exists = df_food[(df_food['Name'] == food_name) & (df_food['Label'] == food_label)]

        if food_exists.empty:
            print(f"Food '{food_name}' with label '{food_label}' not found. Please log the food first.")
            return None
        else:
            # Fix: Use .loc to avoid the copy warning
            food_data = food_exists.iloc[0].copy()
            food_data['Calories'] *= multiplier
            food_data['Protein'] *= multiplier
            food_data['Carbs'] *= multiplier
            food_data['Fat_Saturated'] *= multiplier
            food_data['Fat_Regular'] *= multiplier
            food_data['Sodium'] *= multiplier
            return food_data

    def is_duplicate_meal(self, foods):
        """
        Check if a meal with the same foods and their multipliers already exists in the meal log.
        """
        df_meals = pd.read_excel(self.meal_excel_file)

        if df_meals.empty:
            return False

        # Initialize a list to store all meals as separate lists of foods
        existing_meals = []

        # Iterate over each unique meal name in the existing meal log
        for meal_name in df_meals['Meal_Name'].unique():
            # Get all rows related to the current meal name
            meal_df = df_meals[df_meals['Meal_Name'] == meal_name]

            # Create a list of tuples (food name, label, multiplier) for this specific meal
            meal_foods = [
                (row['Food_Name'].split('x ')[-1], row['Label'], float(row['Food_Name'].split('x ')[0]))
                for idx, row in meal_df.iterrows()
                if 'Total' not in row['Food_Name']  # Skip the 'Total' row
            ]

            # Add the current meal's list of foods as a separate entry in existing_meals
            existing_meals.append(meal_foods)

        # Now, create a list for the current meal being created
        current_meal_foods = [
            (food_name, food_label, multiplier)
            for food_name, food_label, multiplier in foods
        ]

        # Sort the current meal foods list to ensure order doesn't affect comparison
        current_meal_foods.sort()

        # Compare the current meal foods list with each existing meal's foods list
        for existing_meal_foods in existing_meals:
            # Sort existing meal foods for proper comparison
            existing_meal_foods.sort()
            # Compare the current meal with each existing meal
            if existing_meal_foods == current_meal_foods:
                return True  # Duplicate found

        return False  # No duplicate found

    def create_meal(self, foods):
        # Check if the meal is a duplicate
        if self.is_duplicate_meal(foods):
            return f"Meal '{self.meal_name}' is a duplicate and won't be logged."

        
        total_calories = 0
        total_protein = 0
        total_carbs = 0
        total_fat_saturated = 0
        total_fat_regular = 0
        total_sodium = 0

        meal_data = []

        for food_name, food_label, multiplier in foods:
            food_data = self.add_food_to_meal(food_name, food_label, multiplier)
            if food_data is not None:
                # Accumulate macros based on the multiplier
                total_calories += food_data['Calories']
                total_protein += food_data['Protein']
                total_carbs += food_data['Carbs']
                total_fat_saturated += food_data['Fat_Saturated']
                total_fat_regular += food_data['Fat_Regular']
                total_sodium += food_data['Sodium']

                # Append individual food details
                meal_data.append({
                    'Meal_Name': self.meal_name,
                    'Food_Name': f"{multiplier}x {food_data['Name']}",
                    'Label': food_data['Label'],
                    'Measurement': f"{multiplier}x {food_data['Measurement']}",
                    'Calories':food_data['Calories'],
                    'Protein': food_data['Protein'],
                    'Carbs': food_data['Carbs'],
                    'Fat_Saturated': food_data['Fat_Saturated'],
                    'Fat_Regular': food_data['Fat_Regular'],
                    'Sodium': food_data['Sodium']
                })
            else:
                return f"Meal wasnt added due to insufficient indrients listed!"

        # Add total row at the end
        meal_data.append({
            'Meal_Name': self.meal_name,
            'Food_Name': 'Total',
            'Label': '',
            'Measurement': '',
            'Calories': total_calories,
            'Protein': total_protein,
            'Carbs': total_carbs,
            'Fat_Saturated': total_fat_saturated,
            'Fat_Regular': total_fat_regular,
            'Sodium': total_sodium
        })

        # Save the meal data to the meal log and format the summary row
        self.save_meal(meal_data)

    def save_meal(self, meal_data):
        df_meals = pd.read_excel(self.meal_excel_file)

        # Append the meal data to the meal log
        for data in meal_data:
            df_meals = df_meals._append(data, ignore_index=True)

        df_meals.to_excel(self.meal_excel_file, index=False)

        # Load the workbook for styling
        self.style_total_row(len(meal_data))

        print(f"Meal '{self.meal_name}' created successfully with the following foods:")
        for food in meal_data[:-1]:  # Exclude the total row from the print
            print(f"  {food['Food_Name']} - {food['Measurement']}, Calories: {round(food['Calories'], 2)}, Protein: {round(food['Protein'], 2)}, Carbs: {round(food['Carbs'], 2)}, Fats: {round(food['Fat_Saturated'], 2)} Saturated, {round(food['Fat_Regular'], 2)} Regular, Sodium: {round(food['Sodium'], 2)}")

        # Print total macros for the meal
        total_row = meal_data[-1]
        print(f"Total Macros: Calories: {round(total_row['Calories'], 2)}, Protein: {round(total_row['Protein'], 2)}, Carbs: {round(total_row['Carbs'], 2)}, Fats: {round(total_row['Fat_Saturated'], 2)} Saturated, {round(total_row['Fat_Regular'], 2)} Regular, Sodium: {round(total_row['Sodium'], 2)}")

    def style_total_row(self, meal_length):
        # Open the workbook and get the sheet
        wb = load_workbook(self.meal_excel_file)
        ws = wb.active

        # The total row is the last one we just appended
        total_row_idx = ws.max_row

        # Apply green fill to the total row
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        # Apply fill to the entire row (all columns A to I)
        for col in range(1, 10):  # Assuming 9 columns (A to I)
            ws.cell(row=total_row_idx, column=col).fill = green_fill

        # Save the workbook with the style
        wb.save(self.meal_excel_file)