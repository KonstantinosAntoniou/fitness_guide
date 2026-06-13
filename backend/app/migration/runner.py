"""Seed the DB from legacy Excel exports. Idempotent on foods (name+brand)."""
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models import Meal, MealItem
from app.repositories import FoodRepository, MealRepository
from app.migration.parsers import food_from_row, parse_meal_food_name


def _rows(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    header = [str(h) for h in data[0]]
    return [dict(zip(header, r)) for r in data[1:] if any(c is not None for c in r)]


def migrate(session: Session, foods_path: str, meals_path: str | None = None) -> dict:
    foods = FoodRepository(session)
    report = {"foods_added": 0, "meals_added": 0, "meal_items_added": 0, "skipped": 0}

    for row in _rows(foods_path):
        if not row.get("Name"):
            continue
        if foods.find_by_name_brand(str(row["Name"]), str(row.get("Label") or "")):
            continue
        foods.add(food_from_row(row))
        report["foods_added"] += 1
    session.flush()

    if meals_path:
        meals = MealRepository(session)
        grouped: dict[str, list[dict]] = {}
        for row in _rows(meals_path):
            name = row.get("Meal_Name")
            if name:
                grouped.setdefault(str(name), []).append(row)

        for meal_name, items in grouped.items():
            if meals.find_by_name(meal_name):
                continue
            meal = Meal(name=meal_name)
            session.add(meal)
            session.flush()
            for row in items:
                try:
                    mult, food_name = parse_meal_food_name(str(row.get("Food_Name") or ""))
                except ValueError:
                    continue  # Total / summary rows
                food = foods.find_by_name_brand(food_name, str(row.get("Label") or ""))
                if not food:
                    report["skipped"] += 1
                    continue
                session.add(MealItem(meal_id=meal.id, food_id=food.id, servings=mult))
                report["meal_items_added"] += 1
            report["meals_added"] += 1

    return report
